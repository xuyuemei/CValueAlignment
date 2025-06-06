from openpyxl import Workbook
from openpyxl import load_workbook
import jsonlines
from get import *
from prompt import *
import re

file_ok=jsonlines.open("/home/qzh/Value_cn/dataset/extra/deepseek1000.jsonl","a")

workbook = load_workbook('/home/qzh/Value_cn/dataset/extra/extra1000.xlsx')

# 选择第一个工作表
sheet = workbook.active

ok_inx=1
for row in sheet.iter_rows(min_row=737, min_col=1, max_col=1): #row从1开始
    for cell in row:
        print("第"+str(ok_inx)+"条：")
        ok_inx+=1
        text = cell.value
        text=re.sub(r'\s+', '', cell.value)
        print(text)
        result=get_answer(cn_prompt(text))
        if('answer' in result.keys() and result['answer']==None):
            continue
        result["text"]=text         
        # print(result)
        jsonlines.Writer.write(file_ok,result)   


file_ok.close()
