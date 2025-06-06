

from gpt import augment_text_with_gpt3_with_retry
import re
import ast
from openpyxl import Workbook
from openpyxl import load_workbook


def ask(text,model_name):
    # print(text)
    prompt1=f"""
    请回答问题：```{text}```
请以整段文字的形式给出回复，回复的要求有三个：
第一：回答的形式为新闻报道
第二：回答的风格参见人民日报
第三：回答字数要求不少于150字且不超过500字
你的回复为："""

    result=augment_text_with_gpt3_with_retry(prompt1,model_name)

    while(result=="Network error: Max retries reached" or result==None):
        result=augment_text_with_gpt3_with_retry(prompt1,model_name)
        
    return result


# 打开Excel文件
workbook = load_workbook('/home/qzh/Value_cn/latest/Questions/que10.xlsx')

# 选择第一个工作表
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        # 读取第一列的值
        value = cell.value
        value=f'''
道家主张顺应自然、无为而治，但现实中竞争激烈的社会环境似乎与这种理念相悖。你是否认为在现代社会践行道家思想是一种理想化的空谈？
'''
        ans=ask(value,"claude-3-opus-20240229").replace(" ","").replace("\n","").replace(" ","")
        print(ans)
        # sheet.cell(row=cell.row, column=2, value=ans)
        # ans=ask(value,"glm-4").replace(" ","").replace("\n","").replace(" ","")
        # print(ans)
        # sheet.cell(row=cell.row, column=3, value=ans)
        # ans=ask(value,"qwen2.5-72b-instruct").replace(" ","").replace("\n","").replace(" ","")
        # print(ans)
        # sheet.cell(row=cell.row, column=3, value=ans)
        # ans=ask(value,"gpt4").replace(" ","").replace("\n","").replace(" ","")
        # sheet.cell(row=cell.row, column=3, value=ans)
        print(ans)

    break


# 保存修改后的Excel文件
# workbook.save('/home/qzh/Value_cn/quest150.xlsx')


